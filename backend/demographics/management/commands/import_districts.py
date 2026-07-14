from pathlib import Path
from django.core.management.base import BaseCommand, CommandError
from django.utils.text import slugify
from demographics.models import District


class Command(BaseCommand):
    help = "Import or update districts from an Excel (.xlsx) worksheet."

    # Exact header names expected in the supplied workbook.
    required_columns = {"name", "population_total", "population_male", "population_female"}

    def add_arguments(self, parser):
        parser.add_argument("workbook", type=Path)
        parser.add_argument("--sheet", default=None, help="Worksheet name; defaults to the active sheet")

    def handle(self, *args, **options):
        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise CommandError("Install openpyxl first: pip install openpyxl") from error

        path = options["workbook"]
        if not path.exists():
            raise CommandError(f"Workbook not found: {path}")
        workbook = load_workbook(path, data_only=True, read_only=True)
        sheet = workbook[options["sheet"]] if options["sheet"] else workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value else "" for value in next(rows)]
        missing = self.required_columns - set(headers)
        if missing:
            raise CommandError(f"Missing required columns: {', '.join(sorted(missing))}")

        imported = 0
        for row in rows:
            values = dict(zip(headers, row))
            if not values.get("name"):
                continue
            name = str(values.pop("name")).strip()
            clean = {
                key: value for key, value in values.items()
                if key in {field.name for field in District._meta.fields} and value is not None
            }
            District.objects.update_or_create(
                name=name,
                defaults={"slug": slugify(name), **clean},
            )
            imported += 1
        self.stdout.write(self.style.SUCCESS(f"Imported {imported} districts."))
