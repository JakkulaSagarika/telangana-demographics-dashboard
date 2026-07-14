from collections import Counter
from pathlib import Path
from django.core.management.base import BaseCommand, CommandError
from django.utils.text import slugify
from demographics.models import District, LocalBody


def clean_name(value):
    return " ".join(str(value or "").strip().upper().split())


def records_by_district(sheet, district_column, name_column):
    """Count unique named local bodies in the provided source sheet."""
    values = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        district, name = clean_name(row[district_column]), clean_name(row[name_column])
        if district and name:
            values.add((district, name))
    return Counter(district for district, _ in values)


def read_local_bodies(sheet, body_type, birth_columns, death_columns):
    """Merge the birth and death sections of one workbook sheet by body name."""
    records = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        birth_district, birth_name = clean_name(row[birth_columns[0]]), str(row[birth_columns[1]] or "").strip()
        death_district, death_name = clean_name(row[death_columns[0]]), str(row[death_columns[1]] or "").strip()
        if birth_district and birth_name:
            key = (birth_district, clean_name(birth_name))
            records.setdefault(key, {"name": birth_name, "body_type": body_type, "births": 0, "deaths": 0, "births_male": 0, "births_female": 0, "deaths_male": 0, "deaths_female": 0})
            records[key].update({"births": int(row[birth_columns[2]] or 0), "births_male": int(row[birth_columns[3]] or 0), "births_female": int(row[birth_columns[4]] or 0)})
        if death_district and death_name:
            key = (death_district, clean_name(death_name))
            records.setdefault(key, {"name": death_name, "body_type": body_type, "births": 0, "deaths": 0, "births_male": 0, "births_female": 0, "deaths_male": 0, "deaths_female": 0})
            records[key].update({"deaths": int(row[death_columns[2]] or 0), "deaths_male": int(row[death_columns[3]] or 0), "deaths_female": int(row[death_columns[4]] or 0)})
    return records


class Command(BaseCommand):
    help = "Import district birth/death totals and local-body counts from the NIC workbook."

    def add_arguments(self, parser):
        parser.add_argument("workbook", type=Path)

    def handle(self, *args, **options):
        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise CommandError("Install openpyxl first: pip install openpyxl") from error

        path = options["workbook"]
        if not path.exists():
            raise CommandError(f"Workbook not found: {path}")
        book = load_workbook(path, read_only=True, data_only=True)
        required = {"BIRTH COUNT", "DEATH_COUNT", "GRAM PANCHAYAT", "MUNICIPALITY", "MUNICIPAL CORPORATIN"}
        missing = required - set(book.sheetnames)
        if missing:
            raise CommandError(f"Expected sheets missing: {', '.join(sorted(missing))}")

        def summary_totals(sheet_name):
            totals = {}
            for row in book[sheet_name].iter_rows(min_row=2, values_only=True):
                district = clean_name(row[0])
                total = row[1]
                if district and isinstance(total, (int, float)):
                    totals[district] = {
                        "total": int(total),
                        "male": int(row[2] or 0),
                        "female": int(row[3] or 0),
                        "gram_panchayat": int(row[6] or 0),
                        "municipality": int(row[11] or 0),
                        "corporation": int(row[16] or 0),
                    }
            return totals

        births, deaths = summary_totals("BIRTH COUNT"), summary_totals("DEATH_COUNT")
        gram_panchayats = records_by_district(book["GRAM PANCHAYAT"], 0, 1)
        municipalities = records_by_district(book["MUNICIPALITY"], 0, 1)
        corporations = records_by_district(book["MUNICIPAL CORPORATIN"], 0, 1)

        districts = sorted(set(births) | set(deaths))
        for name in districts:
            District.objects.update_or_create(
                name=name.title(),
                defaults={
                    "slug": slugify(name),
                    "births": births.get(name, {}).get("total", 0),
                    "births_male": births.get(name, {}).get("male", 0),
                    "births_female": births.get(name, {}).get("female", 0),
                    "deaths": deaths.get(name, {}).get("total", 0),
                    "deaths_male": deaths.get(name, {}).get("male", 0),
                    "deaths_female": deaths.get(name, {}).get("female", 0),
                    "gram_panchayat_births": births.get(name, {}).get("gram_panchayat", 0),
                    "gram_panchayat_deaths": deaths.get(name, {}).get("gram_panchayat", 0),
                    "municipality_births": births.get(name, {}).get("municipality", 0),
                    "municipality_deaths": deaths.get(name, {}).get("municipality", 0),
                    "corporation_births": births.get(name, {}).get("corporation", 0),
                    "corporation_deaths": deaths.get(name, {}).get("corporation", 0),
                    "gram_panchayats": gram_panchayats.get(name, 0),
                    "municipalities": municipalities.get(name, 0),
                    "municipal_corporations": corporations.get(name, 0),
                    "facts": "Birth and death figures are sourced from the provided NIC workbook.",
                },
            )
        imported_bodies = 0
        source_sheets = [
            ("GRAM PANCHAYAT", LocalBody.BodyType.GRAM_PANCHAYAT, (0, 1, 2, 3, 4), (6, 7, 8, 9, 10)),
            ("MUNICIPALITY", LocalBody.BodyType.MUNICIPALITY, (0, 1, 2, 3, 4), (7, 8, 9, 10, 11)),
            ("MUNICIPAL CORPORATIN", LocalBody.BodyType.CORPORATION, (0, 1, 2, 3, 4), (7, 8, 9, 10, 11)),
        ]
        LocalBody.objects.all().delete()
        districts_by_name = {clean_name(district.name): district for district in District.objects.all()}
        for sheet_name, body_type, birth_columns, death_columns in source_sheets:
            for (district_name, _), values in read_local_bodies(book[sheet_name], body_type, birth_columns, death_columns).items():
                district = districts_by_name.get(district_name)
                if district:
                    LocalBody.objects.create(district=district, **values)
                    imported_bodies += 1
        self.stdout.write(self.style.SUCCESS(f"Imported birth/death data for {len(districts)} districts and {imported_bodies} local bodies."))
